import detection
import clear_result
import cv2
import openpyxl
import os

test_path = './test/'
result_path = './result'
clear_result.clear(result_path)
num = len([lists for lists in os.listdir(test_path) if os.path.isfile(os.path.join(test_path, lists))])

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "record"

header = ['编号','路径','折痕数','折痕位置','暗斑数','暗斑位置','云纹数','云纹位置','缺陷']
i = 0
for each_header in header:
    sheet.cell(row = 1, column = i+1, value = each_header)
    i += 1

cap=cv2.VideoCapture(0)
i=1
while True:
    #从摄像头读取图片
    sucess,img=cap.read()
    #转为灰度图片
    gray=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    #显示摄像头，背景是灰度。
    #cv2.imshow("img",gray)
    cv2.imshow("img",img)
    #保持画面的持续。
    k=cv2.waitKey(1)
    if k == 27:
        #通过esc键退出摄像
        cv2.destroyAllWindows()
        break
    elif k==ord("s"):
        #通过s键保存图片，并退出。
        cv2.imwrite("%i.jpg"%i,img)

        #cv2.destroyAllWindows()s
        #break
        dir_path = str(i) + '.jpg'
        paper = cv2.imread(dir_path)
        sheet.cell(row=i + 1, column=1, value=(str(i) + '.jpg'))
        sheet.cell(row=i + 1, column=2, value=dir_path)
        x, y = paper.shape[0:2]
        r = 800 / max(x, y)
        paper = cv2.resize(paper, (0, 0), fx=r, fy=r, interpolation=cv2.INTER_NEAREST)
        cv2.imshow('original' + str(i), paper)

        paper_result = detection.detect(paper, i, sheet)
        cv2.imshow('result' + str(i), paper_result)

        cv2.waitKey()
        cv2.destroyWindow('original' + str(i))
        cv2.destroyWindow('result' + str(i))
        i = i + 1
    wb.save('./result/record.xlsx')
#关闭摄像头
cap.release()
cv2.destroyAllWindows()

