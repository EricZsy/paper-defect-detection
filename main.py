import detection
import clear_result
import cv2
import openpyxl
import os

test_path = './test/'
result_path = './result'
clear_result.clear(result_path)
num = len([lists for lists in os.listdir(test_path) if os.path.isfile(os.path.join(test_path, lists))])

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "record"

header = ['编号','路径','折痕数','折痕位置','暗斑数','暗斑位置','云纹数','云纹位置','缺陷']
i = 0
for each_header in header:
    sheet.cell(row = 1, column = i+1, value = each_header)
    i += 1

for m in range(1, num+1):
    dir_path = test_path + str(m) + '.jpg'
    paper = cv2.imread(dir_path)
    sheet.cell(row=m+1, column=1, value=(str(m)+'.jpg'))
    sheet.cell(row=m+1, column=2, value=dir_path)
    x, y = paper.shape[0:2]
    r = 800/max(x, y)
    paper = cv2.resize(paper, (0, 0), fx=r, fy=r, interpolation=cv2.INTER_NEAREST)
    cv2.imshow('original'+str(m), paper)

    paper_result = detection.detect(paper, m, sheet)
    cv2.imshow('result'+str(m), paper_result)

    cv2.waitKey()
    cv2.destroyWindow('original'+str(m))
    cv2.destroyWindow('result' + str(m))
wb.save('./result/record.xlsx')


cv2.destroyAllWindows()

